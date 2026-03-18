import json
from dataclasses import asdict
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font

from backend.storage.case_set_repository import SqliteCaseSetRepository
from backend.storage.run_repository import SqliteMetricSetRepository, SqliteRunRepository, SqliteScheduleRepository, SqliteTaskRepository


TASK_REPORT_PROFILES = [
    {
        "profile_id": "task-report-excel",
        "name": "Excel 执行总览",
        "format_type": "excel",
        "description": "适合交付评测结果，包含任务整体概览与用例明细两个页签。",
        "sections": ["任务概览", "用例明细"],
        "file_extension": ".xlsx",
    },
    {
        "profile_id": "task-report-json",
        "name": "JSON 全量快照",
        "format_type": "json",
        "description": "保留任务、执行、用例明细与趋势摘要，适合系统间集成。",
        "sections": ["task", "latest_execution", "case_results", "trend_snapshot"],
        "file_extension": ".json",
    },
]


PROFILE_MAP = {item["profile_id"]: item for item in TASK_REPORT_PROFILES}


def list_task_report_profiles() -> List[dict]:
    return TASK_REPORT_PROFILES


def build_task_report_bundle(
    task_repo: SqliteTaskRepository,
    run_repo: SqliteRunRepository,
    case_set_repo: SqliteCaseSetRepository,
    metric_repo: SqliteMetricSetRepository,
    schedule_repo: SqliteScheduleRepository,
    task_id: str,
    run_id: Optional[str] = None,
) -> Optional[Dict[str, object]]:
    task = task_repo.get(task_id)
    if not task:
        return None
    selected_run_id = run_id or task.latest_execution_id
    execution = run_repo.get(selected_run_id) if selected_run_id else None
    case_results = run_repo.list_case_results(selected_run_id) if selected_run_id else []
    case_set = case_set_repo.get_case_set(task.case_set_id)
    metric_set = metric_repo.get(task.metric_set_id)
    schedule = schedule_repo.get_active_for_task(task.task_id)
    history = run_repo.list_by_task(task.task_id, limit=20)
    accuracy_series = [
        {
            "run_id": item.run_id,
            "started_at": item.started_at,
            "accuracy": item.accuracy,
            "execution_status": item.execution_status,
        }
        for item in history
        if item.total_cases > 0
    ]
    trend_snapshot = {
        "latest_accuracy": execution.accuracy if execution else None,
        "history_points": accuracy_series,
        "delta_vs_previous": None,
    }
    if len(accuracy_series) >= 2:
        trend_snapshot["delta_vs_previous"] = round(accuracy_series[0]["accuracy"] - accuracy_series[1]["accuracy"], 4)

    return {
        "task": task,
        "case_set": case_set,
        "metric_set": metric_set,
        "schedule": schedule,
        "latest_execution": execution,
        "case_results": case_results,
        "execution_history": history,
        "trend_snapshot": trend_snapshot,
    }


def _write_overview_sheet(sheet, bundle: Dict[str, object]) -> None:
    task = bundle["task"]
    case_set = bundle["case_set"]
    metric_set = bundle["metric_set"]
    execution = bundle["latest_execution"]
    schedule = bundle["schedule"]

    sheet.title = "任务概览"
    sheet["A1"] = "任务结果报告"
    sheet["A1"].font = Font(size=16, bold=True)
    rows = [
        ("任务名称", task.name),
        ("任务ID", task.task_id),
        ("用例集", case_set.name if case_set else task.case_set_id),
        ("启动方式", task.launch_mode),
        ("环境", task.environment_id),
        ("指标参数集", metric_set.name if metric_set else task.metric_set_id),
        ("关联定时任务", schedule.name if schedule else "未关联"),
        ("最新执行ID", execution.run_id if execution else "未执行"),
        ("执行状态", execution.execution_status if execution else "未执行"),
        ("总用例数", execution.total_cases if execution else 0),
        ("已执行用例数", execution.executed_cases if execution else 0),
        ("准确率", round((execution.accuracy if execution else 0) * 100, 2)),
        ("开始时间", execution.started_at if execution else "未执行"),
        ("结束时间", execution.ended_at if execution else "未执行"),
    ]
    row_cursor = 3
    for label, value in rows:
        sheet.cell(row=row_cursor, column=1, value=label)
        sheet.cell(row=row_cursor, column=2, value=value)
        row_cursor += 1


def _write_case_sheet(sheet, bundle: Dict[str, object]) -> None:
    sheet.title = "用例明细"
    headers = ["用例ID", "标题", "类型", "准确率", "状态", "问题标签", "摘要"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column_index, value=header).font = Font(bold=True)
    for row_index, item in enumerate(bundle["case_results"], start=2):
        sheet.cell(row=row_index, column=1, value=item.case_id)
        sheet.cell(row=row_index, column=2, value=item.case_title)
        sheet.cell(row=row_index, column=3, value=item.case_type)
        sheet.cell(row=row_index, column=4, value=round(item.accuracy * 100, 2))
        sheet.cell(row=row_index, column=5, value=item.status)
        sheet.cell(row=row_index, column=6, value="、".join(item.issue_tags))
        sheet.cell(row=row_index, column=7, value=item.summary)


def export_task_report_excel(bundle: Dict[str, object]) -> bytes:
    workbook = Workbook()
    overview_sheet = workbook.active
    _write_overview_sheet(overview_sheet, bundle)
    case_sheet = workbook.create_sheet("用例明细")
    _write_case_sheet(case_sheet, bundle)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def export_task_report_json(bundle: Dict[str, object]) -> bytes:
    payload = {
        "task": asdict(bundle["task"]),
        "case_set": asdict(bundle["case_set"]) if bundle["case_set"] else None,
        "metric_set": asdict(bundle["metric_set"]) if bundle["metric_set"] else None,
        "schedule": asdict(bundle["schedule"]) if bundle["schedule"] else None,
        "latest_execution": asdict(bundle["latest_execution"]) if bundle["latest_execution"] else None,
        "case_results": [asdict(item) for item in bundle["case_results"]],
        "execution_history": [asdict(item) for item in bundle["execution_history"]],
        "trend_snapshot": bundle["trend_snapshot"],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def export_task_report(profile_id: str, bundle: Dict[str, object]) -> Tuple[dict, bytes, str]:
    profile = PROFILE_MAP.get(profile_id)
    if not profile:
        raise LookupError("report profile not found")
    if profile_id == "task-report-excel":
        return profile, export_task_report_excel(bundle), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if profile_id == "task-report-json":
        return profile, export_task_report_json(bundle), "application/json"
    raise LookupError("report profile not found")
