import io
import json
import re
from typing import Dict, List

from openpyxl import Workbook, load_workbook


TEMPLATE_COLUMNS = {
    "NL2SQL": ["case_id", "question", "expected_sql"],
    "NL2CHART": ["case_id", "question", "sql", "expected_chart_type"],
    "智能问数": ["case_id", "question", "expected_sql", "expected_chart_type"],
    "报告多轮交互": [
        "case_id",
        "user_goal",
        "template_name",
        "dialogue_script",
        "param_ground_truth",
        "outline_ground_truth",
        "content_assertions",
    ],
}

REPORT_JSON_FIELDS = {
    "dialogue_script",
    "param_ground_truth",
    "outline_ground_truth",
    "content_assertions",
}

REPORT_JSON_DEFAULTS = {
    "dialogue_script": [],
    "param_ground_truth": {},
    "outline_ground_truth": {},
    "content_assertions": [],
}


def _safe_sheet_title(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", name)[:31]


def export_case_set_workbook(case_set_name: str, case_set_type: str, cases: List[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_title(case_set_name) or "Cases"
    columns = TEMPLATE_COLUMNS[case_set_type]
    sheet.append(columns)
    for case in cases:
        payload = case["payload"]
        row = []
        for column in columns:
            value = payload.get(column, "")
            if column in REPORT_JSON_FIELDS and value != "":
                value = json.dumps(value, ensure_ascii=False)
            row.append(value)
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_case_set_workbook(case_set_type: str, content: bytes) -> List[dict]:
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    expected_headers = TEMPLATE_COLUMNS[case_set_type]
    if headers != expected_headers:
        raise ValueError("template mismatch")

    seen_case_ids = set()
    cases = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or all(value in (None, "") for value in row):
            continue
        payload: Dict[str, object] = {}
        for column, value in zip(expected_headers, row):
            if column in REPORT_JSON_FIELDS:
                payload[column] = json.loads(value) if value else REPORT_JSON_DEFAULTS[column]
            else:
                payload[column] = value or ""
        case_id = str(payload.get("case_id", "")).strip()
        if not case_id:
            raise ValueError("case_id is required")
        if case_id in seen_case_ids:
            raise ValueError("case_id must be unique")
        seen_case_ids.add(case_id)
        cases.append(
            {
                "case_id": case_id,
                "title": str(payload.get("question") or payload.get("user_goal") or case_id),
                "payload": payload,
            }
        )
    return cases
