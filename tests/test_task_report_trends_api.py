import os
import sys
import tempfile
import unittest
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, ROOT)

import backend.app as app_module

DB_DIR = os.path.join(ROOT, "backend")


class TaskReportTrendApiTests(unittest.TestCase):
    def setUp(self):
        self.db_paths = []
        self.meta_db = self._make_db_path("report-eval-")
        self.run_db = self._make_db_path("runs-")
        self.case_set_db = self._make_db_path("case-sets-")
        app_module.DEFAULT_META_DB = self.meta_db
        app_module.DEFAULT_RUN_DB = self.run_db
        app_module.DEFAULT_CASE_SET_DB = self.case_set_db

    def tearDown(self):
        for path in self.db_paths:
            if os.path.exists(path):
                os.remove(path)

    def _make_db_path(self, prefix):
        fd, path = tempfile.mkstemp(prefix=prefix, suffix=".db", dir=DB_DIR)
        os.close(fd)
        self.db_paths.append(path)
        return path

    def _create_task(self, client, *, case_set_id="cs-nl2sql", launch_mode="deferred", name="趋势任务"):
        payload = {
            "name": name,
            "case_set_id": case_set_id,
            "environment_id": "env-staging",
            "metric_set_id": "metric-default",
            "repeat_count": 1,
            "launch_mode": launch_mode,
        }
        response = client.post("/api/tasks", json=payload)
        self.assertEqual(response.status_code, 200)
        return response.json()["task"]

    def test_task_report_profiles_and_excel_export(self):
        with TestClient(app_module.app) as client:
            profiles_resp = client.get("/api/task-report-profiles")
            self.assertEqual(profiles_resp.status_code, 200)
            profiles = profiles_resp.json()["profiles"]
            self.assertTrue(any(item["profile_id"] == "task-report-excel" for item in profiles))
            self.assertTrue(any(item["profile_id"] == "task-report-json" for item in profiles))

            task = self._create_task(client, launch_mode="immediate", name="零售导出任务")
            export_resp = client.post(f"/api/tasks/{task['task_id']}/export", json={"profile_id": "task-report-excel"})
            self.assertEqual(export_resp.status_code, 200)
            self.assertEqual(
                export_resp.headers["content-type"],
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            workbook = load_workbook(filename=BytesIO(export_resp.content))
            self.assertIn("任务概览", workbook.sheetnames)
            self.assertIn("用例明细", workbook.sheetnames)
            overview_sheet = workbook["任务概览"]
            detail_sheet = workbook["用例明细"]
            self.assertEqual(overview_sheet["A1"].value, "任务结果报告")
            detail_rows = list(detail_sheet.iter_rows(values_only=True))
            self.assertTrue(any(row and row[0] == "NL2SQL-014" for row in detail_rows[1:]))

            json_resp = client.post(f"/api/tasks/{task['task_id']}/export", json={"profile_id": "task-report-json"})
            self.assertEqual(json_resp.status_code, 200)
            payload = json_resp.json()
            self.assertIn("task", payload)
            self.assertIn("case_results", payload)
            self.assertTrue(payload["case_results"])

    def test_case_set_and_case_trend_analytics(self):
        with TestClient(app_module.app) as client:
            task = self._create_task(client, case_set_id="cs-nl2sql", launch_mode="deferred", name="零售趋势任务")
            task_id = task["task_id"]
            for _ in range(4):
                execute_resp = client.post(f"/api/tasks/{task_id}/execute")
                self.assertEqual(execute_resp.status_code, 200)

            case_set_trend_resp = client.get("/api/case-sets/cs-nl2sql/trends")
            self.assertEqual(case_set_trend_resp.status_code, 200)
            case_set_trends = case_set_trend_resp.json()
            self.assertGreaterEqual(len(case_set_trends["run_series"]), 4)
            self.assertIn("summary", case_set_trends)
            self.assertIn("regression_alerts", case_set_trends)
            self.assertTrue(case_set_trends["regression_alerts"])

            case_trend_resp = client.get("/api/case-sets/cs-nl2sql/cases/NL2SQL-014/trends")
            self.assertEqual(case_trend_resp.status_code, 200)
            case_trend = case_trend_resp.json()
            self.assertEqual(case_trend["case_id"], "NL2SQL-014")
            self.assertGreaterEqual(len(case_trend["accuracy_series"]), 4)
            self.assertIn("latest_delta", case_trend["summary"])

            overview_resp = client.get("/api/analytics/overview")
            self.assertEqual(overview_resp.status_code, 200)
            overview = overview_resp.json()
            self.assertIn("global_accuracy_series", overview)
            self.assertIn("case_set_summaries", overview)
            self.assertIn("regression_alerts", overview)
            self.assertTrue(overview["case_set_summaries"])


if __name__ == "__main__":
    unittest.main()
