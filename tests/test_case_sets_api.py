import io
import os
import sys
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, ROOT)

import backend.app as app_module
from fastapi.testclient import TestClient

DB_DIR = os.path.join(ROOT, 'backend')


class CaseSetApiTests(unittest.TestCase):
    def setUp(self):
        self.db_paths = []
        self.meta_db = self._make_db_path('report-eval-')
        self.run_db = self._make_db_path('runs-')
        self.case_set_db = self._make_db_path('case-sets-')
        app_module.DEFAULT_META_DB = self.meta_db
        app_module.DEFAULT_RUN_DB = self.run_db
        app_module.DEFAULT_CASE_SET_DB = self.case_set_db

    def tearDown(self):
        for path in self.db_paths:
            if os.path.exists(path):
                os.remove(path)

    def _make_db_path(self, prefix):
        fd, path = tempfile.mkstemp(prefix=prefix, suffix='.db', dir=DB_DIR)
        os.close(fd)
        self.db_paths.append(path)
        return path

    def test_list_case_sets(self):
        with TestClient(app_module.app) as client:
            response = client.get('/api/case-sets')
            self.assertEqual(response.status_code, 200)
            data = response.json()
            self.assertIn('case_sets', data)
            self.assertTrue(any(item['id'] == 'cs-nl2sql' for item in data['case_sets']))

    def test_export_case_set_returns_excel(self):
        with TestClient(app_module.app) as client:
            response = client.get('/api/case-sets/cs-nl2sql/export')
            self.assertEqual(response.status_code, 200)
            self.assertEqual(
                response.headers['content-type'],
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            workbook = load_workbook(io.BytesIO(response.content))
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(headers, ['case_id', 'question', 'expected_sql'])
            self.assertEqual(sheet['A2'].value, 'NL2SQL-014')

    def test_import_case_set_overwrites_existing_cases(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['case_id', 'question', 'expected_sql'])
        sheet.append(['NL2SQL-999', '本月 GMV', 'SELECT sum(gmv) FROM sales'])
        content = io.BytesIO()
        workbook.save(content)
        content.seek(0)

        with TestClient(app_module.app) as client:
            response = client.post(
                '/api/case-sets/cs-nl2sql/import',
                files={'file': ('cs-nl2sql.xlsx', content.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
            )
            self.assertEqual(response.status_code, 200)

            detail = client.get('/api/case-sets/cs-nl2sql')
            self.assertEqual(detail.status_code, 200)
            payload = detail.json()
            self.assertEqual(len(payload['cases']), 1)
            self.assertEqual(payload['cases'][0]['case_id'], 'NL2SQL-999')

    def test_import_rejects_invalid_template(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['case_id', 'question', 'wrong_column'])
        content = io.BytesIO()
        workbook.save(content)
        content.seek(0)

        with TestClient(app_module.app) as client:
            response = client.post(
                '/api/case-sets/cs-nl2sql/import',
                files={'file': ('invalid.xlsx', content.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
            )
            self.assertEqual(response.status_code, 400)
            self.assertIn('template', response.json()['detail'])

    def test_seeded_case_payloads_match_case_set_types(self):
        with TestClient(app_module.app) as client:
            nl2sql = client.get('/api/case-sets/cs-nl2sql').json()['cases'][0]['payload']
            self.assertEqual(set(nl2sql.keys()), {'case_id', 'question', 'expected_sql'})

            nl2chart = client.get('/api/case-sets/cs-nl2chart').json()['cases'][0]['payload']
            self.assertEqual(set(nl2chart.keys()), {'case_id', 'question', 'sql', 'expected_chart_type'})

            smart_qa = client.get('/api/case-sets/cs-smart-qa').json()['cases'][0]['payload']
            self.assertEqual(set(smart_qa.keys()), {'case_id', 'question', 'expected_sql', 'expected_chart_type'})

            report = client.get('/api/case-sets/cs-report').json()['cases'][0]['payload']
            self.assertEqual(
                set(report.keys()),
                {
                    'case_id',
                    'user_goal',
                    'template_name',
                    'dialogue_script',
                    'param_ground_truth',
                    'outline_ground_truth',
                    'content_assertions',
                },
            )
            self.assertIsInstance(report['dialogue_script'], list)
            self.assertIsInstance(report['param_ground_truth'], dict)
            self.assertIsInstance(report['outline_ground_truth'], dict)
            self.assertIsInstance(report['content_assertions'], list)


if __name__ == '__main__':
    unittest.main()
